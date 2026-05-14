from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
import pandas as pd
import sqlite3
import io
import os
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

app = FastAPI(title="Warehouse List Processor")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH = "warehouse.db"

# ─── Database Setup ───────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS warehouses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            created_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS quotas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            warehouse_id INTEGER REFERENCES warehouses(id) ON DELETE CASCADE,
            sku TEXT NOT NULL,
            updated_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS quota_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            warehouse_id INTEGER REFERENCES warehouses(id),
            warehouse_name TEXT,
            skus_count INTEGER,
            note TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        );
    """)
    conn.commit()
    conn.close()

init_db()

# ─── Helper: Parse Excel ──────────────────────────────────────────────────────
def parse_excel(file_bytes: bytes):
    """Parse any excel format into a standard dataframe"""
    df_raw = pd.read_excel(io.BytesIO(file_bytes), header=None, sheet_name=None)
    
    # Pick first non-empty sheet
    sheet = None
    for name, df in df_raw.items():
        if len(df) > 2:
            sheet = df
            break
    if sheet is None:
        raise ValueError("الملف فاضي أو مش قادر يقرأه")

    # Find header row
    keywords = ['الصنف', 'الكود', 'السعر', 'الكمية', 'الكميه', 'رقم', 'vendorsku', 'productname']
    header_idx = 0
    for i in range(min(10, len(sheet))):
        row_str = ' '.join(str(c).lower() for c in sheet.iloc[i])
        if any(k.lower() in row_str for k in keywords):
            header_idx = i
            break

    headers = [str(c).strip() for c in sheet.iloc[header_idx]]
    data_rows = sheet.iloc[header_idx + 1:].reset_index(drop=True)
    data_rows.columns = headers

    # Map columns
    col_map = {}
    patterns = {
        'VendorSKU': ['رقم الصنف', 'الكود', 'كود الصنف', 'vendorsku', 'رقم', 'كود'],
        'ProductName': ['إسم الصنف', 'اسم الصنف', 'الصنف', 'الأسم عربي', 'productname'],
        'StockQuantity': ['الكميه', 'الكمية', 'stockquantity', 'quantity'],
        'PriceBeforeDiscount': ['السعر', 'سعر ج', 'سعر الجمهور', 'pricebeforediscount', 'price'],
        'Discount': ['خصم اساسى', 'الخصم', 'خصم %', 'نقدي', 'جملة', 'مندوب', 'discount'],
        'InlineQuota': ['ك الكوته', 'الكوته', 'كوته', 'كوتة', 'ك الكوتة', 'quota'],
    }

    for field, pats in patterns.items():
        for col in headers:
            col_clean = col.lower().replace(' ', '')
            for pat in pats:
                if col_clean == pat.lower().replace(' ', '') or col == pat:
                    col_map[field] = col
                    break
            if field in col_map:
                break

    if 'VendorSKU' not in col_map:
        raise ValueError("مش قادر يلاقي عمود رقم الصنف")

    # Build clean dataframe
    rows = []
    for _, row in data_rows.iterrows():
        sku = str(row.get(col_map.get('VendorSKU', ''), '')).strip()
        if not sku or sku in ('nan', 'رقم الصنف', ''):
            continue
        rows.append({
            'VendorSKU': sku,
            'ProductName': str(row.get(col_map.get('ProductName', ''), '')).strip(),
            'StockQuantity': int(float(str(row.get(col_map.get('StockQuantity', ''), 0)).replace(',', '') or 0)),
            'PriceBeforeDiscount': float(str(row.get(col_map.get('PriceBeforeDiscount', ''), 0)).replace(',', '') or 0),
            'Discount': round(float(str(row.get(col_map.get('Discount', ''), 0)).replace(',', '') or 0)),
            'InlineQuota': int(float(str(row.get(col_map.get('InlineQuota', ''), 0) if 'InlineQuota' in col_map else 0).replace(',', '') or 0)),
        })

    has_inline_quota = 'InlineQuota' in col_map
    return rows, has_inline_quota


def build_output_excel(rows):
    """Build output excel matching the import template"""
    cols = ['VendorSKU', 'StockQuantity', 'PriceBeforeDiscount', 'Discount', 'OrderMaximumQuantity', 'ProductName']
    df = pd.DataFrame(rows, columns=cols)
    
    wb = load_workbook('/app/template.xlsx') if os.path.exists('/app/template.xlsx') else None
    
    output = io.BytesIO()
    if wb:
        ws = wb.active
        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j).value = val
        wb.save(output)
    else:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
    
    output.seek(0)
    return output


# ─── API Routes ───────────────────────────────────────────────────────────────

@app.get("/api/warehouses")
def get_warehouses():
    conn = get_db()
    rows = conn.execute("""
        SELECT w.id, w.name, w.created_at,
               COUNT(q.id) as quota_count,
               MAX(q.updated_at) as last_quota_update
        FROM warehouses w
        LEFT JOIN quotas q ON q.warehouse_id = w.id
        GROUP BY w.id
        ORDER BY w.name
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@app.post("/api/warehouses")
def create_warehouse(data: dict):
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(400, "اسم المخزن مطلوب")
    conn = get_db()
    try:
        conn.execute("INSERT INTO warehouses (name) VALUES (?)", (name,))
        conn.commit()
    except sqlite3.IntegrityError:
        raise HTTPException(400, "المخزن ده موجود بالفعل")
    finally:
        conn.close()
    return {"status": "ok", "message": f"تم إضافة المخزن: {name}"}


@app.delete("/api/warehouses/{warehouse_id}")
def delete_warehouse(warehouse_id: int):
    conn = get_db()
    conn.execute("DELETE FROM warehouses WHERE id = ?", (warehouse_id,))
    conn.commit()
    conn.close()
    return {"status": "ok"}


@app.get("/api/warehouses/{warehouse_id}/quota")
def get_quota(warehouse_id: int):
    conn = get_db()
    rows = conn.execute(
        "SELECT sku, updated_at FROM quotas WHERE warehouse_id = ? ORDER BY sku",
        (warehouse_id,)
    ).fetchall()
    history = conn.execute(
        "SELECT skus_count, note, created_at FROM quota_history WHERE warehouse_id = ? ORDER BY created_at DESC LIMIT 10",
        (warehouse_id,)
    ).fetchall()
    conn.close()
    return {
        "skus": [r["sku"] for r in rows],
        "updated_at": rows[0]["updated_at"] if rows else None,
        "history": [dict(h) for h in history]
    }


@app.post("/api/warehouses/{warehouse_id}/quota")
async def save_quota(
    warehouse_id: int,
    file: UploadFile = File(...),
    note: str = Form("")
):
    content = await file.read()
    rows, _ = parse_excel(content)
    skus = list({r['VendorSKU'] for r in rows})

    conn = get_db()
    warehouse = conn.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
    if not warehouse:
        raise HTTPException(404, "المخزن مش موجود")

    # Save quota history
    conn.execute(
        "INSERT INTO quota_history (warehouse_id, warehouse_name, skus_count, note) VALUES (?, ?, ?, ?)",
        (warehouse_id, warehouse["name"], len(skus), note)
    )
    # Replace quota
    conn.execute("DELETE FROM quotas WHERE warehouse_id = ?", (warehouse_id,))
    now = datetime.now().isoformat()
    for sku in skus:
        conn.execute(
            "INSERT INTO quotas (warehouse_id, sku, updated_at) VALUES (?, ?, ?)",
            (warehouse_id, sku, now)
        )
    conn.commit()
    conn.close()
    return {"status": "ok", "skus_count": len(skus), "message": f"تم حفظ {len(skus)} صنف في الكوتة"}


@app.post("/api/process")
async def process_list(
    warehouse_id: int = Form(...),
    list_file: UploadFile = File(...),
    quota_file: UploadFile = File(None),
):
    # Parse main list
    list_content = await list_file.read()
    rows, has_inline_quota = parse_excel(list_content)

    # Get quota SKUs
    quota_skus = set()

    # 1. From uploaded quota file
    if quota_file and quota_file.filename:
        quota_content = await quota_file.read()
        quota_rows, _ = parse_excel(quota_content)
        for r in quota_rows:
            quota_skus.add(r['VendorSKU'])
        # Save as new quota for this warehouse
        conn = get_db()
        warehouse = conn.execute("SELECT name FROM warehouses WHERE id = ?", (warehouse_id,)).fetchone()
        if warehouse:
            conn.execute(
                "INSERT INTO quota_history (warehouse_id, warehouse_name, skus_count, note) VALUES (?, ?, ?, ?)",
                (warehouse_id, warehouse["name"], len(quota_skus), "رُفع مع الليسته")
            )
            conn.execute("DELETE FROM quotas WHERE warehouse_id = ?", (warehouse_id,))
            now = datetime.now().isoformat()
            for sku in quota_skus:
                conn.execute("INSERT INTO quotas (warehouse_id, sku, updated_at) VALUES (?, ?, ?)", (warehouse_id, sku, now))
            conn.commit()
        conn.close()
    else:
        # 2. Use saved quota from DB
        conn = get_db()
        saved = conn.execute("SELECT sku FROM quotas WHERE warehouse_id = ?", (warehouse_id,)).fetchall()
        quota_skus = {r["sku"] for r in saved}
        conn.close()

    # Build output
    inline_count = 0
    file_quota_count = 0
    out_rows = []

    for row in rows:
        if has_inline_quota and row['InlineQuota'] > 0:
            max_qty = row['InlineQuota']
            inline_count += 1
        elif row['VendorSKU'] in quota_skus:
            max_qty = 1
            file_quota_count += 1
        else:
            max_qty = 10000

        out_rows.append({
            'VendorSKU': row['VendorSKU'],
            'StockQuantity': row['StockQuantity'],
            'PriceBeforeDiscount': row['PriceBeforeDiscount'],
            'Discount': row['Discount'],
            'OrderMaximumQuantity': max_qty,
            'ProductName': row['ProductName'],
        })

    output = build_output_excel(out_rows)
    filename = list_file.filename.replace('.xlsx', '').replace('.xls', '') + '_import.xlsx'

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


# ─── Serve Frontend ───────────────────────────────────────────────────────────
if os.path.exists("static"):
    app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
def serve_frontend():
    return FileResponse("static/index.html")
